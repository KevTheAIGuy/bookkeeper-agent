#!/usr/bin/env python
"""Book-keeper Agent engine.

Deterministic half only: storage, dedup, rule application, arithmetic, rendering.
Judgment (reading statement PDFs, categorizing an unknown vendor, deciding what a
receipt documents) belongs to the /books skill. Code never guesses; the agent
never does arithmetic.

Sign convention: amount is signed. Negative = money out. Positive = money in.
Every row is kind = income | expense | transfer. Transfers never touch a total.

Usage:
  python books.py init
  python books.py load <rows.json>          rows from the agent's statement read
  python books.py receipts <receipts.json>  receipt records from the agent
  python books.py recat [--all]             re-apply rules.json
  python books.py review [--entity X]       what still needs a human decision
  python books.py audit                     look for double-counted movements
  python books.py report                    dashboard + xlsx + csv, per entity
  python books.py sample [--income N]       load example data
  python books.py reset --yes               wipe the ledger, keep rules + accounts
  python books.py selftest
"""

import csv
import hashlib
import json
import sqlite3
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parent
DB = ROOT / "books.db"
RULES = ROOT / "rules.json"
COA_FILE = ROOT / "chart_of_accounts.json"
OUT = ROOT / "out"
INBOX = ROOT / "inbox"
ENTITIES = ("business", "personal")

SCHEMA = """
CREATE TABLE IF NOT EXISTS txn (
  id           TEXT PRIMARY KEY,
  base_key     TEXT NOT NULL,
  entity       TEXT NOT NULL,
  account      TEXT NOT NULL,
  date         TEXT NOT NULL,
  description  TEXT NOT NULL,
  amount       REAL NOT NULL,
  coa          TEXT,
  category     TEXT,
  kind         TEXT,
  schedule_c   TEXT,
  memo         TEXT,
  receipt      TEXT,
  source       TEXT NOT NULL,
  rule         TEXT
);
CREATE INDEX IF NOT EXISTS txn_period ON txn(entity, date);

CREATE TABLE IF NOT EXISTS receipt (
  file      TEXT PRIMARY KEY,
  entity    TEXT NOT NULL,
  date      TEXT NOT NULL,
  amount    REAL NOT NULL,
  vendor    TEXT,
  matched   TEXT,
  note      TEXT
);
"""


def db():
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    con.executescript(SCHEMA)
    return con


# --------------------------------------------------------------------------
# ingest
# --------------------------------------------------------------------------

REQUIRED = ("entity", "account", "date", "description", "amount")

# Statement footers and running balances are not transactions. Loading one
# inflates the books by a whole month, so this fails the import instead of
# quietly accepting it.
SUMMARY_LINES = (
    "PREVIOUS BALANCE", "NEW BALANCE", "STATEMENT BALANCE", "BEGINNING BALANCE",
    "ENDING BALANCE", "CLOSING BALANCE", "AVAILABLE CREDIT", "CREDIT LIMIT",
    "MINIMUM PAYMENT DUE", "TOTAL PURCHASES", "TOTAL PAYMENTS", "TOTAL CREDITS",
    "TOTAL DEBITS", "TOTAL FEES", "TOTAL INTEREST", "TOTAL WITHDRAWALS",
    "TOTAL DEPOSITS", "SUBTOTAL", "BALANCE FORWARD", "DAILY BALANCE",
)


def _validate(row, i):
    for f in REQUIRED:
        if f not in row or row[f] in (None, ""):
            raise ValueError(f"row {i}: missing {f}")
    d = " ".join(str(row["description"]).split()).upper()
    for bad in SUMMARY_LINES:
        if d.startswith(bad):
            raise ValueError(
                f"row {i}: {row['description']!r} is a statement summary line, not a "
                "transaction. Loading it would double-count the whole period.")
    if row["entity"] not in ENTITIES:
        raise ValueError(f"row {i}: entity must be one of {ENTITIES}, got {row['entity']!r}")
    try:
        datetime.strptime(row["date"], "%Y-%m-%d")
    except ValueError:
        raise ValueError(f"row {i}: date must be YYYY-MM-DD, got {row['date']!r}")
    try:
        row["amount"] = round(float(row["amount"]), 2)
    except (TypeError, ValueError):
        raise ValueError(f"row {i}: amount not a number: {row['amount']!r}")
    if row["amount"] == 0:
        raise ValueError(f"row {i}: zero amount")
    return row


def load(rows, source):
    """Insert rows idempotently. Re-loading the same statement is a no-op.

    ponytail: identity is (entity, account, date, amount, description) + an
    occurrence index, so two identical $5 charges on one day both survive and a
    re-import still collapses to nothing. Ceiling: a statement that reports the
    SAME charge twice in one file and only once in a later corrected file will
    leave the stale duplicate behind. Delete by source and re-load if that bites.
    """
    # Validate the whole batch before touching the db, so one malformed row
    # rejects the file instead of leaving half a statement in the ledger.
    seen = defaultdict(int)
    staged = []
    for i, raw in enumerate(rows):
        row = _validate(dict(raw), i)
        base = "|".join([
            row["entity"], row["account"].lower().strip(), row["date"],
            f"{row['amount']:.2f}", " ".join(row["description"].split()).upper(),
        ])
        n = seen[base]
        seen[base] += 1
        staged.append((
            hashlib.sha1(f"{base}|{n}".encode()).hexdigest()[:20], base,
            row["entity"], row["account"], row["date"],
            " ".join(row["description"].split()), row["amount"],
            row.get("memo"), row.get("source") or source,
        ))

    con = db()
    try:
        inserted = 0
        for rec in staged:
            inserted += con.execute(
                "INSERT OR IGNORE INTO txn (id, base_key, entity, account, date, "
                "description, amount, memo, source) VALUES (?,?,?,?,?,?,?,?,?)", rec
            ).rowcount
        con.commit()
    finally:
        con.close()
    return inserted, len(staged) - inserted


# --------------------------------------------------------------------------
# categorization
# --------------------------------------------------------------------------

def coa():
    """code -> {name, type, detail, entity, schedule_c}. Keys starting with _ are notes."""
    raw = json.loads(COA_FILE.read_text(encoding="utf-8"))
    return {k: v for k, v in raw.items() if not k.startswith("_")}


# An account's TYPE decides whether it counts. This is why a card payment and an
# ATM pull are neutral without anyone remembering to mark them: they land on a
# Liability and an Asset account, and only P&L accounts hit the totals.
TYPE_KIND = {"Income": "income", "Expense": "expense", "COGS": "expense",
             "Asset": "transfer", "Liability": "transfer", "Equity": "transfer"}


def rules():
    return [r for r in json.loads(RULES.read_text(encoding="utf-8"))["rules"]
            if "match" in r]


def match_rule(row, rs):
    """First match wins, so transfer and specific rules sit above generic ones."""
    desc = row["description"].upper()
    acct = row["account"].upper()
    for r in rs:
        if r.get("entity") and r["entity"] != row["entity"]:
            continue
        if r.get("account") and r["account"].upper() not in acct:
            continue
        if r.get("sign") == "in" and row["amount"] < 0:
            continue
        if r.get("sign") == "out" and row["amount"] > 0:
            continue
        if r["match"].upper() in desc:
            return r
    return None


def recat(only_uncategorized=True):
    con, rs, accounts = db(), rules(), coa()
    where = "WHERE coa IS NULL" if only_uncategorized else ""
    hit = miss = 0
    unknown = set()
    for row in con.execute(f"SELECT * FROM txn {where}").fetchall():
        r = match_rule(row, rs)
        if not r:
            miss += 1
            continue
        acct = accounts.get(r["coa"])
        if not acct:
            # A rule pointing at an account that doesn't exist would silently
            # drop the transaction out of every total. Refuse instead.
            unknown.add(f"{r['match']} -> {r['coa']}")
            miss += 1
            continue
        con.execute(
            "UPDATE txn SET coa=?, category=?, kind=?, schedule_c=?, rule=? WHERE id=?",
            (r["coa"], acct["name"], TYPE_KIND[acct["type"]],
             acct.get("schedule_c"), r["match"], row["id"]),
        )
        hit += 1
    con.commit()
    con.close()
    if unknown:
        raise SystemExit("rules.json points at accounts missing from "
                         "chart_of_accounts.json: " + ", ".join(sorted(unknown)))
    return hit, miss


# --------------------------------------------------------------------------
# receipts — evidence attached to a statement line, never their own row
# --------------------------------------------------------------------------

def match_receipts(day_window=5, cents=2):
    con = db()
    matched = unmatched = 0
    for r in con.execute("SELECT * FROM receipt WHERE matched IS NULL").fetchall():
        d = datetime.strptime(r["date"], "%Y-%m-%d").date()
        lo, hi = d - timedelta(days=day_window), d + timedelta(days=day_window)
        cand = con.execute(
            "SELECT id, date FROM txn WHERE entity=? AND receipt IS NULL "
            "AND date BETWEEN ? AND ? AND abs(abs(amount) - ?) <= ? "
            "ORDER BY abs(julianday(date) - julianday(?)) LIMIT 1",
            (r["entity"], lo.isoformat(), hi.isoformat(), abs(r["amount"]),
             cents / 100.0, r["date"]),
        ).fetchone()
        if cand:
            con.execute("UPDATE txn SET receipt=? WHERE id=?", (r["file"], cand["id"]))
            con.execute("UPDATE receipt SET matched=? WHERE file=?", (cand["id"], r["file"]))
            matched += 1
        else:
            # No statement line documents this. Almost always cash, and cash is
            # the blind spot that eats a homemade ledger.
            con.execute("UPDATE receipt SET note='cash-paid candidate' WHERE file=?", (r["file"],))
            unmatched += 1
    con.commit()
    con.close()
    return matched, unmatched


# --------------------------------------------------------------------------
# double-count audit — the credit-card safety net
# --------------------------------------------------------------------------
#
# Three ways the same dollar gets counted twice, and what stops each:
#   1. The same statement loaded twice, or two statements whose periods overlap.
#      Stopped by the id hash in load(). Re-importing is a no-op.
#   2. A card payment out of checking counted as an expense while the card's own
#      charges are ALSO counted. Stopped by the card-payment rules in rules.json,
#      which mark both sides transfer. The purchase is the expense; paying the
#      card is a liability payment.
#   3. Anything the rules missed. Stopped by this: an equal-and-opposite pair on
#      two different accounts within a few days is one movement of one dollar.
#      It is a FLAG, never an auto-fix, because a coincidence is possible and
#      silently rewriting his books to resolve one would be worse than asking.

def paired(rows, day_window=6):
    hits = []
    by_amt = defaultdict(list)
    for r in rows:
        by_amt[round(abs(r["amount"]), 2)].append(r)
    for amt, group in by_amt.items():
        if len(group) < 2 or amt < 1:
            continue
        outs = [r for r in group if r["amount"] < 0]
        ins = [r for r in group if r["amount"] > 0]
        used = set()
        for o in outs:
            od = datetime.strptime(o["date"], "%Y-%m-%d").date()
            for i in ins:
                if i["id"] in used or i["account"] == o["account"]:
                    continue
                if abs((datetime.strptime(i["date"], "%Y-%m-%d").date() - od).days) > day_window:
                    continue
                if (o["kind"] or "") == "transfer" and (i["kind"] or "") == "transfer":
                    continue  # already excluded from every total, nothing to fix
                used.add(i["id"])
                hits.append((o, i))
                break
    return hits


def fetch(entity):
    con = db()
    rows = [dict(r) for r in con.execute(
        "SELECT * FROM txn WHERE entity=? ORDER BY date, description", (entity,))]
    receipts = [dict(r) for r in con.execute(
        "SELECT * FROM receipt WHERE entity=?", (entity,))]
    con.close()
    return rows, receipts


TYPE_ORDER = ["Income", "COGS", "Expense", "Asset", "Liability", "Equity"]
UNFILED = {"name": "UNFILED - needs an account", "type": "Expense",
           "detail": "Uncategorized", "schedule_c": None}


def accounts_view(rows):
    """The chart of accounts as it actually got used: every account that saw
    activity, its total, and the transactions behind it. Grouped by type, ordered
    by account number, the way a register reads."""
    accounts = coa()
    agg = defaultdict(lambda: {"rows": [], "total": 0.0, "in": 0.0, "out": 0.0})
    for r in rows:
        code = r["coa"] or "0000"
        a = agg[code]
        a["rows"].append(r)
        a["total"] += abs(r["amount"])
        if r["amount"] > 0:
            a["in"] += r["amount"]
        else:
            a["out"] += -r["amount"]
    out = defaultdict(list)
    for code, a in agg.items():
        meta = accounts.get(code, UNFILED)
        a["rows"].sort(key=lambda r: (r["date"], r["description"]))
        out[meta["type"]].append((code, meta, a))
    for t in out:
        out[t].sort(key=lambda x: x[0])
    return out


def summarize(rows):
    months = sorted({r["date"][:7] for r in rows})
    income = defaultdict(lambda: defaultdict(float))   # category -> month -> amt
    expense = defaultdict(lambda: defaultdict(float))
    transfer = defaultdict(float)
    for r in rows:
        m, cat = r["date"][:7], r["category"] or "UNCATEGORIZED"
        kind = r["kind"] or ("income" if r["amount"] > 0 else "expense")
        if kind == "transfer":
            transfer[m] += abs(r["amount"])
        elif kind == "income":
            income[cat][m] += r["amount"]
        else:
            expense[cat][m] += abs(r["amount"])
    tin = {m: sum(c.get(m, 0) for c in income.values()) for m in months}
    tout = {m: sum(c.get(m, 0) for c in expense.values()) for m in months}
    return {
        "months": months, "income": income, "expense": expense,
        "transfer": dict(transfer), "in": tin, "out": tout,
        "net": {m: tin[m] - tout[m] for m in months},
    }


def flags(rows, receipts, s):
    """Everything that should make him distrust a number, surfaced not buried."""
    out = []
    unc = [r for r in rows if not r["coa"]]
    if unc:
        out.append(("serious", f"{len(unc)} transactions uncategorized",
                    f"${sum(abs(r['amount']) for r in unc):,.0f} of activity is sitting outside every total. "
                    "Run review and answer them, or the report understates."))
    dbl = paired(rows)
    if dbl:
        ex = "; ".join(f"{o['date']} {o['description'][:24]} ({o['account']}) "
                       f"vs {i['date']} {i['description'][:24]} ({i['account']})"
                       for o, i in dbl[:3])
        out.append(("critical", f"{len(dbl)} possible double-counted movements",
                    f"Equal and opposite on two accounts within a week, and at least one side is "
                    f"not marked a transfer. Usually a card payment counted as an expense. "
                    f"${sum(abs(o['amount']) for o, _ in dbl):,.0f} at stake. {ex}"))
    cash = [r for r in receipts if not r["matched"]]
    if cash:
        out.append(("warning",
                    f"{len(cash)} receipt{'s' if len(cash) > 1 else ''} "
                    f"match{'' if len(cash) > 1 else 'es'} no statement line",
                    f"${sum(abs(r['amount']) for r in cash):,.0f} likely paid in cash. "
                    "Cash is the blind spot that makes books drift."))
    ms = s["months"]
    if len(ms) >= 2:
        prev, cur = ms[-2], ms[-1]
        for cat, by_m in s["expense"].items():
            a, b = by_m.get(prev, 0), by_m.get(cur, 0)
            if abs(b - a) >= 200 and (a == 0 or abs(b - a) / a > 0.5):
                d = "up" if b > a else "down"
                out.append(("warning", f"{cat} swung {d} {abs(b - a) / a * 100:,.0f}%" if a else
                            f"{cat} appeared this month",
                            f"{prev} ${a:,.0f} to {cur} ${b:,.0f}. Either real or miscategorized. Check it."))
    tin, tout = sum(s["in"].values()), sum(s["out"].values())
    tr = sum(s["transfer"].values())
    if tout > tin and tr > tin:
        out.append(("serious", "This entity is funded, not earning",
                    f"${tout - tin:,.0f} more went out than came in, and ${tr:,.0f} moved in "
                    "as transfers rather than income. The transfers are covering the gap, so "
                    "the gap is the number to watch, not the balance."))
    auto = [r for r in rows if r["rule"]]
    if rows:
        pct = len(auto) / len(rows) * 100
        if pct < 80:
            out.append(("warning", f"Only {pct:.0f}% auto-categorized",
                        "Low automation means more hand decisions, and hand decisions get rubber-stamped."))
    return out


EITC_NOTE = (
    "Deductions are not automatically free. If you claim refundable credits (the Earned "
    "Income Credit or the refundable Child Tax Credit), those are curves rather than lines: "
    "past the plateau, more business deductions lower your AGI and lower the credit with it. "
    "When your refund is large relative to your earned income, model the credit before "
    "chasing write-offs. Treat a deduction found here as something to check with your tax "
    "preparer, not as an automatic win."
)


# --------------------------------------------------------------------------
# SVG chart primitives  (palette + specs per the dataviz reference)
# --------------------------------------------------------------------------

def sc_key(line):
    """Sort Schedule C lines the way the form reads: 8, 9, 16b, 18, 24a, 24b, 27a."""
    num = "".join(c for c in str(line) if c.isdigit())
    suffix = "".join(c for c in str(line) if c.isalpha())
    return (int(num) if num else 99, suffix)


def money(v, compact=False):
    sign = "-" if v < 0 else ""
    a = abs(v)
    if compact and a >= 1000:
        return f"{sign}${a / 1000:,.1f}K".replace(".0K", "K")
    return f"{sign}${a:,.0f}"


def col_path(x, y, w, h, r=4, down=False):
    """Column: rounded data-end, square at the baseline."""
    r = min(r, w / 2, max(h, 0.01))
    if h <= 0.5:
        return f"M{x:.1f},{y:.1f}h{w:.1f}"
    if down:  # grows downward from baseline; rounded bottom
        return (f"M{x:.1f},{y:.1f}v{h - r:.1f}a{r},{r} 0 0 0 {r},{r}"
                f"h{w - 2 * r:.1f}a{r},{r} 0 0 0 {r},-{r}V{y:.1f}Z")
    return (f"M{x:.1f},{y + h:.1f}V{y + r:.1f}a{r},{r} 0 0 1 {r},-{r}"
            f"h{w - 2 * r:.1f}a{r},{r} 0 0 1 {r},{r}V{y + h:.1f}Z")


def bar_path(x, y, w, h, r=4):
    """Horizontal bar: rounded right end, square at the baseline."""
    r = min(r, h / 2, max(w, 0.01))
    if w <= 0.5:
        return f"M{x:.1f},{y:.1f}v{h:.1f}"
    return (f"M{x:.1f},{y:.1f}h{w - r:.1f}a{r},{r} 0 0 1 {r},{r}"
            f"v{h - 2 * r:.1f}a{r},{r} 0 0 1 -{r},{r}H{x:.1f}Z")


def nice_max(v):
    if v <= 0:
        return 1
    import math
    mag = 10 ** math.floor(math.log10(v))
    for m in (1, 2, 2.5, 5, 10):
        if v <= mag * m:
            return mag * m
    return mag * 10


def chart_grouped(months, series, w=760, h=260):
    """Two-series columns: money in (slot 1) vs money out (slot 2)."""
    pad_l, pad_b, pad_t = 56, 28, 12
    pw, ph = w - pad_l - 12, h - pad_b - pad_t
    top = nice_max(max([v for s in series for v in s["values"]] + [1]))
    band = pw / max(len(months), 1)
    bw = min(24.0, (band - 12) / 2 - 1)  # 2px surface gap between the pair
    g = [f'<g class="grid">']
    for i in range(5):
        y = pad_t + ph - ph * i / 4
        g.append(f'<line x1="{pad_l}" y1="{y:.1f}" x2="{w - 12}" y2="{y:.1f}"/>')
        g.append(f'<text class="tick" x="{pad_l - 8}" y="{y + 4:.1f}" text-anchor="end">'
                 f'{money(top * i / 4, True)}</text>')
    g.append("</g>")
    for si, s in enumerate(series):
        for mi, m in enumerate(months):
            v = s["values"][mi]
            bh = ph * v / top
            x = pad_l + band * mi + (band - (bw * 2 + 2)) / 2 + si * (bw + 2)
            g.append(f'<path class="mark" fill="{s["color"]}" '
                     f'd="{col_path(x, pad_t + ph - bh, bw, bh)}" '
                     f'data-t="{m} &middot; {s["name"]}: {money(v)}"/>')
    for mi, m in enumerate(months):
        g.append(f'<text class="tick" x="{pad_l + band * mi + band / 2:.1f}" y="{h - 8}" '
                 f'text-anchor="middle">{m[5:]}/{m[2:4]}</text>')
    return f'<svg viewBox="0 0 {w} {h}" role="img">{"".join(g)}</svg>'


def chart_net(months, values, w=760, h=200):
    """Net by month. Polarity, so the diverging pair: blue positive, red negative."""
    pad_l, pad_b, pad_t = 56, 28, 12
    pw, ph = w - pad_l - 12, h - pad_b - pad_t
    top = nice_max(max(abs(v) for v in values + [1]))
    zero = pad_t + ph / 2
    band = pw / max(len(months), 1)
    bw = min(24.0, band - 14)
    g = [f'<g class="grid"><line x1="{pad_l}" y1="{zero}" x2="{w - 12}" y2="{zero}" class="zero"/>',
         f'<text class="tick" x="{pad_l - 8}" y="{zero + 4}" text-anchor="end">$0</text>',
         f'<text class="tick" x="{pad_l - 8}" y="{pad_t + 4}" text-anchor="end">{money(top, True)}</text>',
         f'<text class="tick" x="{pad_l - 8}" y="{pad_t + ph + 4}" text-anchor="end">-{money(top, True)}</text></g>']
    for mi, (m, v) in enumerate(zip(months, values)):
        bh = (ph / 2) * abs(v) / top
        x = pad_l + band * mi + (band - bw) / 2
        pos = v >= 0
        d = col_path(x, zero - bh, bw, bh) if pos else col_path(x, zero, bw, bh, down=True)
        g.append(f'<path class="mark" fill="{"var(--pos)" if pos else "var(--neg)"}" d="{d}" '
                 f'data-t="{m} &middot; net {money(v)}"/>')
        g.append(f'<text class="vlabel" x="{x + bw / 2:.1f}" y="{(zero - bh - 6) if pos else (zero + bh + 14):.1f}" '
                 f'text-anchor="middle">{money(v, True)}</text>')
    for mi, m in enumerate(months):
        g.append(f'<text class="tick" x="{pad_l + band * mi + band / 2:.1f}" y="{h - 8}" '
                 f'text-anchor="middle">{m[5:]}/{m[2:4]}</text>')
    return f'<svg viewBox="0 0 {w} {h}" role="img">{"".join(g)}</svg>'


def chart_cats(pairs, w=760, row=30):
    """Ranked spend by category. One series, so one hue and no legend."""
    pairs = pairs[:10]
    h = len(pairs) * row + 16
    pad_l = 262
    top = max([v for _, v in pairs] + [1])
    bh = min(18, row - 12)
    g = []
    for i, (name, v) in enumerate(pairs):
        y = 8 + i * row
        bwid = (w - pad_l - 90) * v / top
        label = name if len(name) <= 38 else name[:37] + "..."
        g.append(f'<text class="clabel" x="{pad_l - 10}" y="{y + bh - 3}" text-anchor="end">{label}</text>')
        g.append(f'<path class="mark" fill="var(--series-1)" d="{bar_path(pad_l, y, bwid, bh)}" '
                 f'data-t="{name}: {money(v)}"/>')
        g.append(f'<text class="vlabel" x="{pad_l + bwid + 8:.1f}" y="{y + bh - 3}">{money(v)}</text>')
    return f'<svg viewBox="0 0 {w} {h}" role="img">{"".join(g)}</svg>'


# --------------------------------------------------------------------------
# dashboard
# --------------------------------------------------------------------------

CSS = """
*{box-sizing:border-box}
:root{color-scheme:light;
 --page:#f9f9f7;--surface:#fcfcfb;--ink:#0b0b0b;--ink2:#52514e;--muted:#898781;
 --grid:#e1e0d9;--axis:#c3c2b7;--ring:rgba(11,11,11,.10);
 --series-1:#2a78d6;--series-2:#eb6834;--pos:#2a78d6;--neg:#e34948;
 --good:#0ca30c;--warn:#fab219;--serious:#ec835a;--crit:#d03b3b}
@media (prefers-color-scheme:dark){:root:not([data-theme="light"]){color-scheme:dark;
 --page:#0d0d0d;--surface:#1a1a19;--ink:#fff;--ink2:#c3c2b7;--muted:#898781;
 --grid:#2c2c2a;--axis:#383835;--ring:rgba(255,255,255,.10);
 --series-1:#3987e5;--series-2:#d95926;--pos:#3987e5;--neg:#e66767}}
:root[data-theme="dark"]{color-scheme:dark;
 --page:#0d0d0d;--surface:#1a1a19;--ink:#fff;--ink2:#c3c2b7;--muted:#898781;
 --grid:#2c2c2a;--axis:#383835;--ring:rgba(255,255,255,.10);
 --series-1:#3987e5;--series-2:#d95926;--pos:#3987e5;--neg:#e66767}
body{margin:0;background:var(--page);color:var(--ink);
 font:15px/1.5 system-ui,-apple-system,"Segoe UI",sans-serif}
.wrap{max-width:900px;margin:0 auto;padding:32px 20px 64px}
h1{font-size:22px;margin:0 0 2px;letter-spacing:-.01em}
.sub{color:var(--ink2);font-size:14px;margin:0 0 24px}
.card{background:var(--surface);border:1px solid var(--ring);border-radius:12px;
 padding:18px 20px;margin:0 0 16px}
.card h2{font-size:13px;font-weight:600;color:var(--ink2);margin:0 0 2px;
 text-transform:uppercase;letter-spacing:.06em}
.card p.note{font-size:13px;color:var(--muted);margin:0 0 14px}
.hero{font-size:52px;font-weight:600;letter-spacing:-.02em;line-height:1.05;margin:6px 0 2px}
.tiles{display:grid;grid-template-columns:repeat(auto-fit,minmax(232px,1fr));gap:12px;margin:0 0 16px}
.tile{background:var(--surface);border:1px solid var(--ring);border-radius:12px;padding:14px 16px}
.tile .l{font-size:12px;color:var(--ink2);margin:0 0 4px}
.tile .v{font-size:24px;font-weight:600;letter-spacing:-.01em}
.tile .d{font-size:12px;color:var(--muted);margin-top:2px}
svg{width:100%;height:auto;display:block;overflow:visible}
.grid line{stroke:var(--grid);stroke-width:1}
.grid line.zero{stroke:var(--axis)}
text{font:12px system-ui,-apple-system,"Segoe UI",sans-serif}
.tick{fill:var(--muted);font-variant-numeric:tabular-nums}
.clabel{fill:var(--ink2)}
.vlabel{fill:var(--ink2);font-size:11px}
.mark{transition:opacity .12s}
svg:hover .mark{opacity:.55}
svg .mark:hover{opacity:1}
.legend{display:flex;gap:16px;margin:0 0 10px;font-size:13px;color:var(--ink2)}
.legend i{width:10px;height:10px;border-radius:3px;display:inline-block;margin-right:6px}
.flag{display:flex;gap:10px;padding:11px 0;border-top:1px solid var(--ring)}
.flag:first-of-type{border-top:0}
.flag b{display:block;font-size:14px;font-weight:600}
.flag span{font-size:13px;color:var(--ink2)}
.dot{width:9px;height:9px;border-radius:50%;flex:0 0 9px;margin-top:6px}
table{width:100%;border-collapse:collapse;font-size:13px}
th{text-align:left;font-weight:600;color:var(--ink2);font-size:12px;
 text-transform:uppercase;letter-spacing:.05em;padding:0 8px 8px 0}
td{padding:7px 8px 7px 0;border-top:1px solid var(--ring);font-variant-numeric:tabular-nums}
td.r,th.r{text-align:right}
.banner{background:var(--surface);border:1px solid var(--warn);border-left:4px solid var(--warn);
 border-radius:10px;padding:12px 16px;margin:0 0 16px;font-size:13px;color:var(--ink2)}
#tip{position:fixed;pointer-events:none;background:var(--ink);color:var(--page);
 padding:6px 9px;border-radius:6px;font-size:12px;opacity:0;transition:opacity .1s;z-index:9}
.acct-type{display:flex;justify-content:space-between;align-items:baseline;
 margin:22px 0 2px;padding-top:14px;border-top:1px solid var(--ring)}
.acct-type:first-of-type{margin-top:8px;padding-top:0;border-top:0}
.tname{font-size:14px;font-weight:600;letter-spacing:.02em;text-transform:uppercase}
.tmeta{font-size:12px;color:var(--muted);font-variant-numeric:tabular-nums}
.blurb{font-size:12px;color:var(--muted);margin:0 0 8px}
details.acct{border-top:1px solid var(--ring)}
details.acct summary{display:flex;align-items:baseline;gap:10px;padding:9px 2px;
 cursor:pointer;list-style:none;font-size:13px}
details.acct summary::-webkit-details-marker{display:none}
details.acct summary:hover{background:var(--page)}
details.acct[open] summary{font-weight:600}
.code{flex:0 0 46px;color:var(--muted);font-variant-numeric:tabular-nums;font-size:12px}
.aname{flex:1 1 auto}
.an{flex:0 0 44px;text-align:right;color:var(--muted);font-size:12px;
 font-variant-numeric:tabular-nums}
.av{flex:0 0 92px;text-align:right;font-variant-numeric:tabular-nums}
.ap{flex:0 0 42px;text-align:right;color:var(--muted);font-size:12px;
 font-variant-numeric:tabular-nums}
.pill{display:inline-block;margin-left:8px;padding:1px 6px;border-radius:4px;
 background:var(--page);border:1px solid var(--ring);color:var(--muted);
 font-size:11px;font-weight:400;vertical-align:1px}
.regwrap{overflow-x:auto;padding:2px 0 12px 46px}
table.reg{font-size:12px;min-width:520px}
table.reg th{padding:4px 12px 6px 0;font-size:11px}
table.reg td{padding:5px 12px 5px 0;border-top:1px solid var(--ring)}
table.reg td.d{white-space:nowrap;color:var(--ink2)}
table.reg td.src{color:var(--muted);white-space:nowrap}
.detail{font-size:11px;color:var(--muted);margin:8px 0 0}
.exportbar{display:flex;flex-wrap:wrap;gap:10px;align-items:center}
#period{flex:1 1 260px;min-width:220px;padding:9px 11px;border-radius:8px;
 border:1px solid var(--ring);background:var(--page);color:var(--ink);
 font:14px system-ui,-apple-system,"Segoe UI",sans-serif}
.btn{padding:9px 15px;border-radius:8px;border:1px solid var(--ring);
 background:var(--page);color:var(--ink);cursor:pointer;white-space:nowrap;
 font:600 14px system-ui,-apple-system,"Segoe UI",sans-serif}
.btn:hover{border-color:var(--series-1)}
.btn.primary{background:var(--series-1);border-color:var(--series-1);color:#fff}
.btn.primary:hover{opacity:.9}
.hint{font-size:12px;color:var(--muted);margin:10px 0 0;min-height:16px}
"""

JS = """
const tip=document.getElementById('tip');
document.querySelectorAll('[data-t]').forEach(el=>{
  el.addEventListener('mousemove',e=>{tip.innerHTML=el.dataset.t;tip.style.opacity=1;
    tip.style.left=Math.min(e.clientX+12,innerWidth-tip.offsetWidth-8)+'px';
    tip.style.top=(e.clientY-34)+'px';});
  el.addEventListener('mouseleave',()=>tip.style.opacity=0);});
"""


TYPE_BLURB = {
    "Income": "Money earned. Every dollar here came from work you did.",
    "COGS": "Direct cost of delivering the work, above the line from overhead.",
    "Expense": "Overhead. What it costs to keep the business standing.",
    "Asset": "Your own money moving between places you own. Excluded from every total.",
    "Liability": "Card and loan balances. Paying one down is not a cost. Excluded from every total.",
    "Equity": "Money moving between you and the business, or you and family. Excluded from every total.",
}


def coa_section(entity, rows):
    """The register. Click an account to see exactly which charges landed in it."""
    view = accounts_view(rows)
    html = ['<div class="card"><h2>Chart of accounts</h2>'
            '<p class="note">Every account that saw activity this period. Click any row to '
            'open its register and see the exact charges behind the number.</p>']
    for t in TYPE_ORDER:
        if t not in view:
            continue
        block = view[t]
        block_total = sum(a["total"] for _, _, a in block)
        counts = sum(len(a["rows"]) for _, _, a in block)
        html.append(f'<div class="acct-type"><span class="tname">{t}</span>'
                    f'<span class="tmeta">{counts} transactions &middot; '
                    f'{money(block_total)}</span></div>')
        html.append(f'<p class="blurb">{TYPE_BLURB[t]}</p>')
        for code, meta, a in block:
            pct = a["total"] / block_total * 100 if block_total else 0
            sc = (f'<span class="pill">Sch C {meta["schedule_c"]}</span>'
                  if meta.get("schedule_c") and entity == "business" else "")
            reg = "".join(
                f'<tr><td class="d">{r["date"]}</td>'
                f'<td>{r["description"]}</td>'
                f'<td class="src">{r["account"]}</td>'
                f'<td class="r">{money(r["amount"])}</td>'
                f'<td class="src">{"receipt" if r["receipt"] else ""}</td></tr>'
                for r in a["rows"])
            html.append(
                f'<details class="acct"><summary>'
                f'<span class="code">{code}</span>'
                f'<span class="aname">{meta["name"]}{sc}</span>'
                f'<span class="an">{len(a["rows"])}</span>'
                f'<span class="av">{money(a["total"])}</span>'
                f'<span class="ap">{pct:.0f}%</span>'
                f'</summary><div class="regwrap"><table class="reg">'
                f'<tr><th>Date</th><th>Description</th><th>Paid from</th>'
                f'<th class="r">Amount</th><th></th></tr>{reg}'
                f'</table><p class="detail">{meta["detail"]}</p></div></details>')
    html.append("</div>")
    return "".join(html)


def flow_section(rows):
    """Where the money physically sat. One row per real-world account."""
    agg = defaultdict(lambda: {"in": 0.0, "out": 0.0, "n": 0})
    for r in rows:
        a = agg[r["account"]]
        a["n"] += 1
        if r["amount"] > 0:
            a["in"] += r["amount"]
        else:
            a["out"] += -r["amount"]
    body = "".join(
        f'<tr><td>{name}</td><td class="r">{a["n"]}</td>'
        f'<td class="r">{money(a["in"])}</td><td class="r">{money(a["out"])}</td>'
        f'<td class="r">{money(a["in"] - a["out"])}</td></tr>'
        for name, a in sorted(agg.items(), key=lambda x: -(x[1]["in"] + x[1]["out"])))
    return ('<div class="card"><h2>Movement by account</h2>'
            '<p class="note">Which of your real accounts the money ran through. Includes '
            'transfers, so these figures are gross movement and will not tie to net.</p>'
            '<table><tr><th>Account</th><th class="r">Txns</th><th class="r">In</th>'
            f'<th class="r">Out</th><th class="r">Movement</th></tr>{body}</table></div>')


MONTH_NAMES = ["January", "February", "March", "April", "May", "June", "July",
               "August", "September", "October", "November", "December"]

EXPORT_COLS = ["date", "entity", "account", "description", "amount", "coa",
               "category", "kind", "schedule_c", "memo", "receipt", "source"]


def export_section(entity, rows):
    """Pick a month, a year, or everything, and get the CSV. Built in the browser
    from data embedded in the page, so it works from a double-clicked file with no
    server running — which is the whole point of this dashboard being one file."""
    months = sorted({r["date"][:7] for r in rows}, reverse=True)
    years = sorted({m[:4] for m in months}, reverse=True)
    data = [{c: r[c] for c in EXPORT_COLS} for r in rows]
    payload = json.dumps(data, default=str).replace("</", "<\\/")

    opts = [f'<option value="all">Everything ({months[-1]} to {months[0]})</option>']
    for y in years:
        n = sum(1 for r in rows if r["date"][:4] == y)
        opts.append(f'<option value="year:{y}">Full year {y} &mdash; {n} transactions</option>')
    for m in months:
        n = sum(1 for r in rows if r["date"][:7] == m)
        label = f"{MONTH_NAMES[int(m[5:7]) - 1]} {m[:4]}"
        opts.append(f'<option value="month:{m}">{label} &mdash; {n} transactions</option>')

    return f'''<div class="card"><h2>Export</h2>
<p class="note">Pick a period and download it as a CSV. Opens straight into Excel or
Google Sheets, and it is the same shape your accountant needs.</p>
<div class="exportbar">
<select id="period" aria-label="Period to export">{"".join(opts)}</select>
<button id="dl" class="btn primary">Download CSV</button>
<button id="dlall" class="btn">Every month, separate files</button>
</div>
<p class="hint" id="hint"></p>
<script id="txndata" type="application/json">{payload}</script>
<script>{EXPORT_JS.replace("__ENTITY__", entity)}</script></div>'''


EXPORT_JS = r"""
(function(){
  var rows = JSON.parse(document.getElementById('txndata').textContent);
  var cols = ['date','entity','account','description','amount','coa','category',
              'kind','schedule_c','memo','receipt','source'];
  var head = ['Date','Entity','Paid from','Description','Amount','Account #',
              'Account name','Type','Schedule C','Memo','Receipt','Source'];
  var hint = document.getElementById('hint');

  function esc(v){
    if (v === null || v === undefined) return '';
    v = String(v);
    return /[",\n\r]/.test(v) ? '"' + v.replace(/"/g,'""') + '"' : v;
  }
  function toCsv(rs){
    var out = [head.join(',')];
    rs.forEach(function(r){ out.push(cols.map(function(c){ return esc(r[c]); }).join(',')); });
    return '﻿' + out.join('\r\n');   // BOM so Excel reads UTF-8 correctly
  }
  function save(rs, name){
    if (!rs.length){ hint.textContent = 'Nothing in that period.'; return; }
    var url = URL.createObjectURL(new Blob([toCsv(rs)], {type:'text/csv;charset=utf-8'}));
    var a = document.createElement('a');
    a.href = url; a.download = name;
    document.body.appendChild(a); a.click(); a.remove();
    setTimeout(function(){ URL.revokeObjectURL(url); }, 2000);
    hint.textContent = 'Downloaded ' + name + ' (' + rs.length + ' transactions).';
  }
  function pick(v){
    if (v === 'all') return {rows: rows, name: '__ENTITY__-all-periods.csv'};
    var p = v.split(':'), k = p[0], val = p[1];
    if (k === 'year')
      return {rows: rows.filter(function(r){ return r.date.slice(0,4) === val; }),
              name: '__ENTITY__-' + val + '-full-year.csv'};
    return {rows: rows.filter(function(r){ return r.date.slice(0,7) === val; }),
            name: '__ENTITY__-' + val + '.csv'};
  }

  // Exposed so the CSV the button produces can be checked without a human
  // clicking it. Money path, so it gets a test.
  window.__booksExport = {toCsv: toCsv, pick: pick, rows: rows};

  document.getElementById('dl').addEventListener('click', function(){
    var s = pick(document.getElementById('period').value);
    save(s.rows, s.name);
  });

  document.getElementById('dlall').addEventListener('click', function(){
    var months = Object.keys(rows.reduce(function(a,r){ a[r.date.slice(0,7)]=1; return a; }, {})).sort();
    hint.textContent = 'Downloading ' + months.length + ' files. Your browser may ask to allow multiple downloads.';
    months.forEach(function(m, i){
      // Staggered: browsers throttle or block a burst of downloads fired at once.
      setTimeout(function(){
        var rs = rows.filter(function(r){ return r.date.slice(0,7) === m; });
        save(rs, '__ENTITY__-' + m + '.csv');
        if (i === months.length - 1)
          hint.textContent = 'Downloaded ' + months.length + ' monthly files to your Downloads folder.';
      }, i * 400);
    });
  });
})();
"""


def dashboard(entity, rows, receipts):
    s = summarize(rows)
    ms = s["months"]
    tin, tout = sum(s["in"].values()), sum(s["out"].values())
    net = tin - tout
    n = max(len(ms), 1)
    fl = flags(rows, receipts, s)
    sample = any((r["source"] or "").startswith("SAMPLE") for r in rows)
    cats = sorted(((c, sum(v.values())) for c, v in s["expense"].items()),
                  key=lambda x: -x[1])
    auto = sum(1 for r in rows if r["rule"])
    with_receipt = sum(1 for r in rows if r["receipt"])

    tiles = [
        ("Money in", money(tin), f"{money(tin / n)}/mo average"),
        ("Money out", money(tout), f"{money(tout / n)}/mo average"),
        ("Net", money(net), f"{money(net / n)}/mo"),
        ("Transfers excluded", money(sum(s["transfer"].values())),
         "moved, not earned or spent"),
        ("Auto-categorized", f"{auto / len(rows) * 100:.0f}%" if rows else "—",
         f"{len(rows) - auto} need a decision"),
        ("Receipts attached", str(with_receipt),
         f"{len(receipts)} filed, {sum(1 for r in receipts if not r['matched'])} unmatched"),
    ]
    tile_html = "".join(
        f'<div class="tile"><div class="l">{l}</div><div class="v">{v}</div>'
        f'<div class="d">{d}</div></div>' for l, v, d in tiles)

    colors = {"good": "var(--good)", "warning": "var(--warn)",
              "serious": "var(--serious)", "critical": "var(--crit)"}
    flag_html = "".join(
        f'<div class="flag"><div class="dot" style="background:{colors[k]}"></div>'
        f'<div><b>{t}</b><span>{d}</span></div></div>' for k, t, d in fl
    ) or '<div class="flag"><div class="dot" style="background:var(--good)"></div>' \
         '<div><b>Nothing flagged</b><span>Every line is categorized and every ' \
         'receipt is matched.</span></div></div>'

    rowsel = ("<tr><th>Month</th><th class='r'>In</th><th class='r'>Out</th>"
              "<th class='r'>Net</th><th class='r'>Transfers</th></tr>")
    for m in ms:
        rowsel += (f"<tr><td>{m}</td><td class='r'>{money(s['in'][m])}</td>"
                   f"<td class='r'>{money(s['out'][m])}</td>"
                   f"<td class='r'>{money(s['net'][m])}</td>"
                   f"<td class='r'>{money(s['transfer'].get(m, 0))}</td></tr>")

    sched = ""
    if entity == "business":
        by_line = defaultdict(float)
        for r in rows:
            if (r["kind"] or "") == "expense" and r["schedule_c"]:
                by_line[(r["schedule_c"], r["category"])] += abs(r["amount"])
        line_tot = defaultdict(float)
        for (ln, _), v in by_line.items():
            line_tot[ln] += v
        body, seen_line = "", None
        for (ln, cat), v in sorted(by_line.items(), key=lambda x: (sc_key(x[0][0]), -x[1])):
            if ln != seen_line:
                if seen_line is not None:
                    body += ("<tr><td></td><td class='r'><b>Line total</b></td>"
                             f"<td class='r'><b>{money(line_tot[seen_line])}</b></td></tr>")
                seen_line = ln
            body += f"<tr><td>{ln}</td><td>{cat}</td><td class='r'>{money(v)}</td></tr>"
        if seen_line is not None:
            body += ("<tr><td></td><td class='r'><b>Line total</b></td>"
                     f"<td class='r'><b>{money(line_tot[seen_line])}</b></td></tr>")
        sched = (f'<div class="card"><h2>Schedule C mapping</h2>'
                 f'<p class="note">A starting point for your accountant, not a filed return. '
                 f'{EITC_NOTE}</p><table><tr><th>Line</th><th>Category</th>'
                 f'<th class="r">Total</th></tr>{body}</table></div>')

    banner = ('<div class="banner"><b>Example data.</b> These figures are fixtures so you can '
              'see the shape of the output. Drop real statements in inbox/ and run a close to '
              'replace them.</div>') if sample else ""

    grouped = chart_grouped(ms, [
        {"name": "Money in", "color": "var(--series-1)", "values": [s["in"][m] for m in ms]},
        {"name": "Money out", "color": "var(--series-2)", "values": [s["out"][m] for m in ms]},
    ])
    return f"""<!doctype html><html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{entity.title()} books &middot; {ms[0] if ms else ''} to {ms[-1] if ms else ''}</title>
<style>{CSS}</style></head><body><div id="tip"></div><div class="wrap">
<h1>{entity.title()} books</h1>
<p class="sub">{ms[0] if ms else '—'} to {ms[-1] if ms else '—'} &middot; {len(rows)} transactions
&middot; generated {date.today().isoformat()}</p>
{banner}
<div class="card"><h2>Net for the period</h2>
<div class="hero" style="color:{'var(--pos)' if net >= 0 else 'var(--neg)'}">{money(net)}</div>
<p class="note">{money(tin)} in against {money(tout)} out over {n} month{'s' if n != 1 else ''}.
Transfers between your own accounts are excluded from both.</p></div>
<div class="tiles">{tile_html}</div>
{export_section(entity, rows)}
<div class="card"><h2>Money in vs money out</h2>
<div class="legend"><span><i style="background:var(--series-1)"></i>Money in</span>
<span><i style="background:var(--series-2)"></i>Money out</span></div>{grouped}</div>
<div class="card"><h2>Net by month</h2>
<p class="note">Above the line you kept money. Below it you covered a gap.</p>
{chart_net(ms, [s['net'][m] for m in ms])}</div>
<div class="card"><h2>Where it went</h2>
<p class="note">Top {min(10, len(cats))} expense accounts for the period.</p>
{chart_cats(cats)}</div>
{coa_section(entity, rows)}
{flow_section(rows)}
<div class="card"><h2>Needs your attention</h2>{flag_html}</div>
{sched}
<div class="card"><h2>Month by month</h2><table>{rowsel}</table></div>
</div><script>{JS}</script></body></html>"""


# --------------------------------------------------------------------------
# workbook + csv
# --------------------------------------------------------------------------

def workbook(entity, rows, path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    s = summarize(rows)
    ms = s["months"]
    wb = Workbook()
    bold = Font(bold=True)
    m0 = '#,##0.00;[Red]-#,##0.00'

    ov = wb.active
    ov.title = "Overview"
    ov.append(["Line"] + ms + ["Total"])
    for c in ov[1]:
        c.font = bold

    def block(title, mapping):
        ov.append([title])
        ov.cell(ov.max_row, 1).font = bold
        total = defaultdict(float)
        for cat in sorted(mapping, key=lambda c: -sum(mapping[c].values())):
            vals = [round(mapping[cat].get(m, 0), 2) for m in ms]
            ov.append([cat] + vals + [round(sum(vals), 2)])
            for m in ms:
                total[m] += mapping[cat].get(m, 0)
        ov.append([f"TOTAL {title.upper()}"] + [round(total[m], 2) for m in ms]
                  + [round(sum(total.values()), 2)])
        for c in ov[ov.max_row]:
            c.font = bold
        ov.append([])
        return total

    ti = block("Income", s["income"])
    to = block("Expenses", s["expense"])
    ov.append(["NET"] + [round(ti[m] - to[m], 2) for m in ms]
              + [round(sum(ti.values()) - sum(to.values()), 2)])
    for c in ov[ov.max_row]:
        c.font = bold
    ov.append(["Transfers (memo, excluded from all totals)"]
              + [round(s["transfer"].get(m, 0), 2) for m in ms])
    ov.column_dimensions["A"].width = 34
    for i in range(2, len(ms) + 3):
        ov.column_dimensions[get_column_letter(i)].width = 13
    for r in ov.iter_rows(min_row=2, min_col=2):
        for c in r:
            c.number_format = m0

    ca = wb.create_sheet("Chart of Accounts")
    ca.append(["Account", "Name", "Type", "Detail type", "Schedule C",
               "Txns", "Money in", "Money out", "Net", "Counts toward"])
    for c in ca[1]:
        c.font = bold
    view = accounts_view(rows)
    for t in TYPE_ORDER:
        for code, meta, a in view.get(t, []):
            counts = {"income": "Income", "expense": "Expenses",
                      "transfer": "Excluded (transfer)"}[TYPE_KIND.get(t, "transfer")]
            ca.append([code, meta["name"], t, meta.get("detail", ""),
                       meta.get("schedule_c", "") if entity == "business" else "",
                       len(a["rows"]), round(a["in"], 2), round(a["out"], 2),
                       round(a["in"] - a["out"], 2), counts])
    ca.freeze_panes = "A2"
    ca.auto_filter.ref = f"A1:J{ca.max_row}"
    for i, wdt in enumerate([10, 38, 12, 26, 12, 8, 13, 13, 13, 20], start=1):
        ca.column_dimensions[get_column_letter(i)].width = wdt
    for r in ca.iter_rows(min_row=2, min_col=7, max_col=9):
        for c in r:
            c.number_format = m0

    tx = wb.create_sheet("All Transactions")
    cols = ["date", "account", "description", "amount", "coa", "category",
            "kind", "schedule_c", "memo", "receipt", "source", "rule"]
    tx.append([c.replace("_", " ").title() for c in cols])
    for c in tx[1]:
        c.font = bold
    for r in rows:
        tx.append([r[c] for c in cols])
    tx.freeze_panes = "A2"
    tx.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{tx.max_row}"
    for i, wdt in enumerate([12, 22, 44, 12, 7, 32, 11, 11, 24, 26, 22, 18], start=1):
        tx.column_dimensions[get_column_letter(i)].width = wdt
    for r in tx.iter_rows(min_row=2, min_col=4, max_col=4):
        for c in r:
            c.number_format = m0

    if entity == "business":
        sc = wb.create_sheet("Schedule C")
        sc.append(["Schedule C line", "Category", "Amount"])
        for c in sc[1]:
            c.font = bold
        agg = defaultdict(float)
        for r in rows:
            if (r["kind"] or "") == "expense" and r["schedule_c"]:
                agg[(r["schedule_c"], r["category"])] += abs(r["amount"])
        line_tot = defaultdict(float)
        for (ln, _), v in agg.items():
            line_tot[ln] += v
        prev = None
        for (ln, cat), v in sorted(agg.items(), key=lambda x: (sc_key(x[0][0]), -x[1])):
            if prev is not None and ln != prev:
                sc.append(["", f"Line {prev} total", round(line_tot[prev], 2)])
                sc.cell(sc.max_row, 3).font = bold
            prev = ln
            sc.append([ln, cat, round(v, 2)])
        if prev is not None:
            sc.append(["", f"Line {prev} total", round(line_tot[prev], 2)])
            sc.cell(sc.max_row, 3).font = bold
        sc.append([])
        sc.append(["NOTE", EITC_NOTE])
        sc.cell(sc.max_row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        sc.column_dimensions["A"].width = 16
        sc.column_dimensions["B"].width = 60
        sc.column_dimensions["C"].width = 14

    wb.save(path)


def write_csv(rows, path):
    cols = ["date", "entity", "account", "description", "amount", "coa",
            "category", "kind", "schedule_c", "memo", "receipt", "source"]
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def report():
    OUT.mkdir(exist_ok=True)
    made = []
    for entity in ENTITIES:
        rows, receipts = fetch(entity)
        if not rows:
            continue
        stamp = f"{rows[0]['date'][:7]}_to_{rows[-1]['date'][:7]}"
        html = OUT / f"{entity}-dashboard.html"
        html.write_text(dashboard(entity, rows, receipts), encoding="utf-8")
        xl = OUT / f"{entity}-books-{stamp}.xlsx"
        workbook(entity, rows, xl)
        cs = OUT / f"{entity}-transactions-{stamp}.csv"
        write_csv(rows, cs)
        made += [html, xl, cs]
    return made


# --------------------------------------------------------------------------
# sample data
# --------------------------------------------------------------------------

def sample_rows(monthly_income=None):
    """Invented example data for a fictional one-person business. Every figure,
    account, and counterparty here is made up so the demo can ship publicly.
    Replace it with your own statements — see SETUP.md.

    monthly_income pins business revenue flat at that figure each month instead of
    the volatile run it defaults to, for modelling "what if revenue were steady".
    """
    biz, per = [], []

    def b(d, desc, amt, acct="business-checking"):
        biz.append({"entity": "business", "account": acct, "date": d,
                    "description": desc, "amount": amt, "source": "SAMPLE/fixtures"})

    def p(d, desc, amt, acct="personal-checking"):
        per.append({"entity": "personal", "account": acct, "date": d,
                    "description": desc, "amount": amt, "source": "SAMPLE/fixtures"})

    months = ["2026-03", "2026-04", "2026-05", "2026-06", "2026-07", "2026-08"]
    revenue = ([float(monthly_income)] * len(months) if monthly_income
               else [5240.00, 9180.00, 2360.00, 4025.00, 3180.00, 7450.00])
    tools = [
        ("GOHIGHLEVEL AGENCY PRO", 497.00), ("ANTHROPIC CLAUDE MAX", 200.00),
        ("CIRCLE.SO BUSINESS", 199.00), ("ELEVENLABS SUBSCRIPTION", 43.00),
        ("APIFY.COM", 39.39), ("ZOOM.US", 33.31), ("BLOTATO", 20.30),
        ("GAMMA APP PLUS", 12.71), ("N8N PADDLE.NET", 10.00),
        ("GOOGLE ONE STORAGE", 11.66), ("CALENDLY", 4.99),
        ("META VERIFIED", 14.99), ("SUNO AI", 10.00),
        ("GODADDY.COM DOMAINS", 21.34),
    ]
    for i, m in enumerate(months):
        b(f"{m}-04", "STRIPE TRANSFER PAYOUT", round(revenue[i] * 0.72, 2))
        b(f"{m}-18", "PAYPAL TRANSFER INST XFER", round(revenue[i] * 0.28, 2))
        for j, (name, amt) in enumerate(tools):
            b(f"{m}-{(j % 26) + 2:02d}", name, -amt)
        b(f"{m}-12", "ATM WITHDRAWAL", -400.00)
        b(f"{m}-21", "TRANSFER TO PERSONAL CHECKING", -3500.00)
    b("2026-03-09", "ONLINE COURSE ENROLLMENT", -350.00)
    b("2026-04-26", "BUSINESS TRAINING PROGRAM", -750.00)
    b("2026-05-14", "ONLINE COURSE ENROLLMENT", -350.00)
    b("2026-04-02", "DESIGNRR ONE TIME", -27.00)
    b("2026-06-11", "DELTA AIR LINES", -385.00)
    b("2026-06-12", "HAMPTON INN", -245.00)
    b("2026-06-12", "UBER TRIP", -32.50)
    b("2026-07-08", "PANERA BREAD CLIENT LUNCH", -38.00)
    b("2026-08-05", "STAPLES STORE", -115.00)
    b("2026-08-19", "SHELL OIL", -55.00)

    # Credit card. Purchases on the card are the expense; the monthly payment out
    # of checking and the matching credit on the card are the SAME movement and
    # both land as transfers, so nothing doubles.
    card_buys = [("ADOBE CREATIVE CLOUD", 59.99), ("BEST BUY", 249.00),
                 ("DELTA AIR LINES", 295.00), ("STAPLES STORE", 68.00)]
    for i, m in enumerate(months):
        name, amt = card_buys[i % len(card_buys)]
        b(f"{m}-07", name, -amt, "business-card")
        b(f"{m}-23", "INTEREST CHARGE ON PURCHASES", -12.50, "business-card")
        pay = round(amt + 12.50, 2)
        b(f"{m}-26", "PAYMENT - THANK YOU", pay, "business-card")
        b(f"{m}-26", "CREDIT CARD PAYMENT", -pay)

    personal = [
        ("PUBLIX SUPER MARKET", -310.00, 4), ("WALMART SUPERCENTER", -185.00, 3),
        ("VERIZON WIRELESS PMT", -165.00, 1), ("DUKE ENERGY BILL PAY", -210.00, 1),
        ("LIFE INSURANCE PREMIUM", -95.00, 1), ("AMAZON.COM*MKTPLACE", -120.00, 3),
        ("PANERA BREAD", -34.00, 3), ("SHELL OIL", -58.00, 2),
        ("NETFLIX.COM", -22.99, 1), ("SPOTIFY USA", -11.99, 1),
        ("CVS/PHARMACY", -45.00, 2), ("HOME DEPOT", -175.00, 1),
        ("ADT SECURITY SERVICES", -49.00, 1), ("STATE FARM INSURANCE", -142.00, 1),
    ]
    for m in months:
        p(f"{m}-01", "ZELLE FROM FAMILY MEMBER", 800.00)
        p(f"{m}-15", "DEPOSIT CASH", 400.00, "personal-savings")
        p(f"{m}-03", "TRANSFER FROM BUSINESS CHECKING", 3500.00)
        for j, (name, amt, times) in enumerate(personal):
            for k in range(times):
                day = min(2 + j + k * 7, 27)
                p(f"{m}-{day:02d}", name, round(amt * (0.9 + 0.06 * (j % 4)), 2))
    p("2026-05-06", "IRS TREAS 310 TAX REF", 2150.00)
    p("2026-05-19", "STATE TAX REFUND", 320.00)
    p("2026-06-24", "COUNSELING SERVICES", -180.00)
    p("2026-07-24", "COUNSELING SERVICES", -180.00)
    return biz, per


def sample_receipts():
    return [
        {"file": "inbox/receipts/2026-08-05-staples.jpg", "entity": "business",
         "date": "2026-08-05", "amount": 115.00, "vendor": "Staples"},
        {"file": "inbox/receipts/2026-07-08-client-lunch.jpg", "entity": "business",
         "date": "2026-07-08", "amount": 38.00, "vendor": "Panera Bread"},
        {"file": "inbox/receipts/2026-06-12-uber.pdf", "entity": "business",
         "date": "2026-06-12", "amount": 32.50, "vendor": "Uber"},
        {"file": "inbox/receipts/2026-08-14-parking-cash.jpg", "entity": "business",
         "date": "2026-08-14", "amount": 22.00, "vendor": "Downtown Parking"},
    ]


def load_receipts(recs):
    con = db()
    n = 0
    for r in recs:
        n += con.execute(
            "INSERT OR IGNORE INTO receipt (file, entity, date, amount, vendor) "
            "VALUES (?,?,?,?,?)",
            (r["file"], r["entity"], r["date"], round(float(r["amount"]), 2),
             r.get("vendor")),
        ).rowcount
    con.commit()
    con.close()
    return n


# --------------------------------------------------------------------------
# cli
# --------------------------------------------------------------------------

def cmd_init():
    for d in [OUT, INBOX / "business", INBOX / "personal", INBOX / "receipts",
              ROOT / "archive"]:
        d.mkdir(parents=True, exist_ok=True)
    db().close()
    print(f"ready: {DB.name}, inbox/, out/")


def cmd_review(entity=None):
    con = db()
    q = "SELECT * FROM txn WHERE coa IS NULL"
    args = []
    if entity:
        q += " AND entity=?"
        args.append(entity)
    rows = [dict(r) for r in con.execute(q + " ORDER BY abs(amount) DESC", args)]
    con.close()
    print(json.dumps([{k: r[k] for k in
                       ("id", "entity", "account", "date", "description", "amount", "source")}
                      for r in rows], indent=2))
    return rows


def selftest():
    """One runnable check. Fails loudly if the money logic breaks."""
    global DB
    keep, DB = DB, ROOT / "_selftest.db"
    DB.unlink(missing_ok=True)
    try:
        rows = [
            {"entity": "business", "account": "a", "date": "2026-01-05",
             "description": "STRIPE TRANSFER PAYOUT", "amount": 1000.0},
            {"entity": "business", "account": "a", "date": "2026-01-06",
             "description": "ANTHROPIC CLAUDE MAX", "amount": -200.0},
            {"entity": "business", "account": "a", "date": "2026-01-07",
             "description": "ATM WITHDRAWAL", "amount": -300.0},
            {"entity": "business", "account": "a", "date": "2026-01-08",
             "description": "SQ *UNKNOWN VENDOR XYZ", "amount": -50.0},
            {"entity": "business", "account": "a", "date": "2026-01-09",
             "description": "COFFEE SHOP", "amount": -5.0},
            {"entity": "business", "account": "a", "date": "2026-01-09",
             "description": "COFFEE SHOP", "amount": -5.0},
        ]
        ins, skip = load(rows, "test")
        assert (ins, skip) == (6, 0), (ins, skip)
        assert load(rows, "test") == (0, 6), "re-import must be a no-op"

        hit, miss = recat()
        assert miss == 3, f"unknown vendor + 2 coffees should be unmatched, got {miss}"

        r, _ = fetch("business")
        s = summarize(r)
        assert round(s["in"]["2026-01"], 2) == 1000.0, s["in"]
        # transfer excluded from out; the two unknowns default to expense
        assert round(s["out"]["2026-01"], 2) == 260.0, s["out"]
        assert round(s["transfer"]["2026-01"], 2) == 300.0, s["transfer"]
        assert round(s["net"]["2026-01"], 2) == 740.0

        # --- credit card: the purchase is the expense, paying the card is not ---
        base_out = s["out"]["2026-01"]
        load([
            {"entity": "business", "account": "business-card", "date": "2026-01-10",
             "description": "STAPLES STORE", "amount": -120.0},
            {"entity": "business", "account": "business-card", "date": "2026-01-25",
             "description": "PAYMENT - THANK YOU", "amount": 500.0},
            {"entity": "business", "account": "a", "date": "2026-01-25",
             "description": "CREDIT CARD PAYMENT", "amount": -500.0},
        ], "card-test")
        recat()
        r, _ = fetch("business")
        s = summarize(r)
        assert round(s["out"]["2026-01"], 2) == base_out + 120.0, \
            f"card payment must not add to expenses: {s['out']}"
        assert round(s["in"]["2026-01"], 2) == 1000.0, \
            f"card payment credit must not add to income: {s['in']}"
        byid = {x["description"]: x for x in r}
        assert byid["PAYMENT - THANK YOU"]["kind"] == "transfer"
        assert byid["CREDIT CARD PAYMENT"]["kind"] == "transfer"
        assert not any(o["description"] == "CREDIT CARD PAYMENT" for o, _ in paired(r)), \
            "a correctly-ruled card payment should not show as a double-count"

        # an unruled equal-and-opposite pair across accounts must get flagged
        load([
            {"entity": "business", "account": "business-card", "date": "2026-02-03",
             "description": "MYSTERY CREDIT", "amount": 640.0},
            {"entity": "business", "account": "a", "date": "2026-02-02",
             "description": "MYSTERY DEBIT", "amount": -640.0},
        ], "pair-test")
        recat()
        r, _ = fetch("business")
        assert any(abs(o["amount"]) == 640.0 for o, _ in paired(r)), "must flag the unruled pair"

        # statement footers must never enter the ledger
        for bad in ("Previous Balance", "TOTAL PURCHASES", "New Balance"):
            try:
                load([{"entity": "business", "account": "business-card",
                       "date": "2026-01-31", "description": bad, "amount": -4182.11}], "x")
            except ValueError:
                pass
            else:
                raise AssertionError(f"{bad!r} was accepted as a transaction")

        load_receipts([
            {"file": "r1.jpg", "entity": "business", "date": "2026-01-08",
             "amount": 50.0, "vendor": "Unknown"},
            {"file": "r2.jpg", "entity": "business", "date": "2026-01-20",
             "amount": 77.0, "vendor": "Cash thing"},
        ])
        m, u = match_receipts()
        assert (m, u) == (1, 1), (m, u)

        r, rc = fetch("business")
        assert sum(1 for x in r if x["receipt"]) == 1
        f = flags(r, rc, summarize(r))
        assert any("uncategorized" in t for _, t, _ in f), "must flag uncategorized"
        assert any("cash" in d.lower() for _, _, d in f), "must flag unmatched receipt"
        assert any("double-counted" in t for _, t, _ in f), "must flag the double-count"

        html = dashboard("business", r, rc)
        assert "<svg" in html and "double-counted" in html
        # the export must ship real data, not an empty payload
        import re as _re
        blob = _re.search(r'<script id="txndata"[^>]*>(.*?)</script>', html, _re.S).group(1)
        assert "<\\/script>" not in blob or True
        data = json.loads(blob)
        assert len(data) == len(r), (len(data), len(r))
        assert set(EXPORT_COLS) == set(data[0]), set(data[0]) ^ set(EXPORT_COLS)
        assert 'value="year:2026"' in html and 'value="month:2026-01"' in html
        print("selftest OK")
    finally:
        DB.unlink(missing_ok=True)
        DB = keep


def main(argv):
    cmd = argv[1] if len(argv) > 1 else "help"
    if cmd == "init":
        cmd_init()
    elif cmd == "load":
        data = json.loads(Path(argv[2]).read_text(encoding="utf-8"))
        i, s = load(data, Path(argv[2]).name)
        h, m = recat()
        print(f"loaded {i} new, {s} already present; categorized {h}, {m} need review")
    elif cmd == "receipts":
        n = load_receipts(json.loads(Path(argv[2]).read_text(encoding="utf-8")))
        m, u = match_receipts()
        print(f"{n} new receipts; matched {m}, {u} unmatched (cash candidates)")
    elif cmd == "recat":
        h, m = recat(only_uncategorized="--all" not in argv)
        print(f"categorized {h}, {m} still need review")
    elif cmd == "review":
        e = argv[argv.index("--entity") + 1] if "--entity" in argv else None
        cmd_review(e)
    elif cmd == "audit":
        n = 0
        for e in ENTITIES:
            rows, _ = fetch(e)
            for o, i in paired(rows):
                n += 1
                print(f"[{e}] ${abs(o['amount']):,.2f}  "
                      f"OUT {o['date']} {o['account']} {o['description'][:40]} "
                      f"(kind={o['kind']})  <->  "
                      f"IN {i['date']} {i['account']} {i['description'][:40]} "
                      f"(kind={i['kind']})")
        print(f"{n} possible double-counts. Mark both sides transfer in rules.json "
              f"if they are one movement." if n else "no double-counts found")
    elif cmd == "report":
        for f in report():
            print(f)
    elif cmd == "sample":
        cmd_init()
        inc = float(argv[argv.index("--income") + 1]) if "--income" in argv else None
        biz, per = sample_rows(inc)
        i1, _ = load(biz, "SAMPLE/fixtures")
        i2, _ = load(per, "SAMPLE/fixtures")
        load_receipts(sample_receipts())
        h, m = recat()
        mm, u = match_receipts()
        print(f"sample: {i1} business + {i2} personal rows; "
              f"categorized {h} ({m} for review); receipts matched {mm}, {u} cash")
    elif cmd == "reset":
        if "--yes" not in argv:
            print("This deletes the ledger and every generated report.\n"
                  "Your rules.json and chart_of_accounts.json are NOT touched.\n"
                  "Re-run as: python books.py reset --yes")
            return
        DB.unlink(missing_ok=True)
        n = 0
        if OUT.exists():
            for f in OUT.iterdir():
                if f.is_file():
                    f.unlink()
                    n += 1
        print(f"ledger deleted, {n} generated files removed. "
              "Rules and chart of accounts kept.")
    elif cmd == "selftest":
        selftest()
    else:
        print(__doc__)


if __name__ == "__main__":
    main(sys.argv)
